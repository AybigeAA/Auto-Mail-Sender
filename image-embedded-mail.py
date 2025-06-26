import os
import base64
import time
import openpyxl
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Gmail API için izinler
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

# Dosya yolları
EXCEL_FILE = "filename.xlsx"
LOGO_TOP = "logo_top.jpg"
LOGO_BOTTOM = "logo_bottom.png"

def get_base64_image(image_path):
    """Resim dosyasını base64 kodlu string olarak döndürür"""
    if not os.path.exists(image_path):
        return ""
    
    with open(image_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
    
    # Dosya uzantısına göre MIME tipini belirle
    mime_type = "image/jpeg" if image_path.lower().endswith(('.jpg', '.jpeg')) else "image/png"
    return f"data:{mime_type};base64,{encoded_string}"

def create_html_template():
  return"""

YOUR MESSAGE


"""


def authenticate_gmail():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def send_email(service, to_email, subject, name, html_template):
    msg = MIMEMultipart('alternative')
    msg['To'] = to_email
    msg['Subject'] = subject

    html_message = html_template   #html_message = html_template.format(name=name)  to send messages with names
    msg.attach(MIMEText(html_message, 'html'))

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    message = {'raw': raw_message}

    try:
        service.users().messages().send(userId='me', body=message).execute()
        print(f"E-posta gönderildi: {to_email}")
    except Exception as error:
        print(f"E-posta gönderimi başarısız oldu: {error}")

def send_emails_from_excel(subject):
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel dosyası bulunamadı: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    service = authenticate_gmail()
    
    # HTML şablonunu bir kere oluştur (resimler base64 olarak içinde)
    html_template = create_html_template()

    # Loop through each row, assuming the first column contains emails
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        email = row[2]  # 3th column for emails
    
        if email:
            send_email(service, email, subject, "", html_template)  # Send email without name
            time.sleep(60)  # Wait for 2 seconds before sending the next email

'''
 to send messages with names get this part of the script out of the comment
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        email = row[0] # 4. sütun
        name = row[3]  # 3. sütun

        if email and name:
            send_email(service, email, subject, name, html_template)
            time.sleep(60)  # Yavaş gönderim için bekleme
'''


if __name__ == "__main__":
    subject = input("Subject of the message ")
    send_emails_from_excel(subject)
