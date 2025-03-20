import os
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
import base64
import openpyxl
import time  # Bekleme süresi için gerekli

# Gmail API için izinler
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

# Mesajı buraya yazın
MESSAGE_TEXT = """



"""
# Excel dosyasının adı sabit
EXCEL_FILE = "filename.xlsx"

def authenticate_gmail():
    """Gmail API'ye bağlanmak için kimlik doğrulama yapar."""
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def send_email(service, to_email, subject, message_text):
    """Gmail API kullanarak e-posta gönderir."""
    message = {
        'raw': base64.urlsafe_b64encode(
            f"To: {to_email}\nSubject: {subject}\n\n{message_text}".encode("utf-8")
        ).decode("utf-8")
    }
    try:
        sent_message = service.users().messages().send(userId='me', body=message).execute()
        print(f"E-posta gönderildi: {to_email}")
        return sent_message
    except Exception as error:
        print(f"E-posta gönderimi başarısız oldu: {error}")

def send_emails_from_excel(subject):
    """Excel dosyasından e-postaları okur ve gönderir."""
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel dosyası bulunamadı: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    service = authenticate_gmail()


    for row in sheet.iter_rows(min_row=2, values_only=True):  # Başlık satırını atla
        email = row[0]  # Sadece ilk sütun, e-posta adresini alır
        if email:
            # Mesajda ad ve soyadı yer almadığı için standart bir mesaj kullanabilirsiniz
            personalized_message = MESSAGE_TEXT  # Önceden tanımlı bir mesaj metni
            send_email(service, email, subject, personalized_message)
        
            # Her e-posta gönderiminden sonra bekleme süresi
            time.sleep(2)  # 2 saniye bekleme
"""
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Başlık satırını atla
        email, name, surname = row[:3]  # İlk üç sütun sırasıyla Email, Name, Surname
        if email :
            personalized_message = MESSAGE_TEXT.format(name=name, surname=surname)
            send_email(service, email, subject, personalized_message)
            
            # Her e-posta gönderiminden sonra bekleme süresi
            time.sleep(2)  # 2 saniye bekleme
"""
if __name__ == "__main__":
    subject = input("E-postaların konusu: ")
    send_emails_from_excel(subject)

